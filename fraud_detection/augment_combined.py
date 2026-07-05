"""
augment_combined.py — 策略4：三策略组合（Trust + Urgency + Emotional）
========================================================================
对 test.csv 中每条对话依次叠加三种诱导策略：
  策略1 建立信任 → 策略2 制造紧迫感 → 策略3 情感操纵

改写顺序符合真实诈骗的渐进逻辑：先建立可信身份，再施加时间压力，
最后用情感诉求推动目标行动。

完成后生成对比 Excel（含颜色标记）。

输出：
  test_combined.csv           — 三策略组合增强测试集
  test_comparison.xlsx         — 代表性样本对比表（含颜色标记）
"""

import os
import re
import random
import pandas as pd

random.seed(42)

# ========================= 共享基础函数 =========================

def parse_dialogue(text):
    if pd.isna(text) or not isinstance(text, str):
        return []
    turns = []
    for line in text.strip().split("\n"):
        line = line.strip()
        if line.startswith("left:"):
            turns.append(("left", line[5:].strip()))
        elif line.startswith("right:"):
            turns.append(("right", line[6:].strip()))
    return turns

def rebuild_dialogue(turns):
    return "\n".join(f"{sp}: {ct}" for sp, ct in turns)

# ========================= 领域检测 =========================

DOMAIN_KEYWORDS = {
    "banking":      ["银行", "贷款", "信用", "利率", "理财", "金融", "放款", "审批",
                     "存款", "流水", "转账", "银保监", "信用卡", "资金"],
    "ecommerce":    ["淘宝", "某宝", "京东", "天猫", "拼多多", "订单", "商品",
                     "退款", "退货", "卖家", "买家", "购物", "收货", "发货"],
    "delivery":     ["快递", "包裹", "物流", "速递", "速达", "配送", "签收", "丢失"],
    "investment":   ["投资", "基金", "股票", "收益", "数字货币", "年化", "回报",
                     "分红", "高收益", "稳健", "风控"],
    "telecom":      ["电讯", "电信", "移动", "联通", "通讯", "话费", "套餐", "宽带"],
    "lottery":      ["彩票", "中奖", "幸运", "大奖", "福彩", "体彩", "奖金"],
    "kidnapping":   ["绑架", "赎金", "人质", "扣押", "保释金", "公安局"],
    "identity":     ["身份证", "盗用", "冒用", "涉案", "洗钱", "通缉", "传票", "检察院",
                     "社保卡", "社保局", "医保", "公积金"],
}

def detect_domain(text):
    if pd.isna(text) or not isinstance(text, str):
        return "generic"
    scores = {d: sum(1 for kw in kws if kw in text) for d, kws in DOMAIN_KEYWORDS.items()}
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "generic"

# ========================= 策略1：建立信任（内联核心逻辑） =========================

def _find_intro_turn(turns):
    for i, (sp, ct) in enumerate(turns):
        if sp != "left":
            continue
        if ct in ["喂", "喂，你好", "你好"]:
            continue
        if ct.strip().endswith("？") and re.search(r'^(?:请问)?是[\u4e00-\u9fa5]+吗', ct):
            continue
        if ct.strip().endswith("?") and re.search(r'^(?:请问)?是[\u4e00-\u9fa5]+吗', ct):
            continue
        if re.search(r'(?:我是|这边是|这里是)', ct):
            return i, ct
    for i, (sp, ct) in enumerate(turns):
        if sp == "left" and len(ct) > 5:
            return i, ct
    return -1, None

def _extract_org(ct, domain, all_turns):
    gov_match = re.search(
        r'(?:我是|这边是|这里是)\s*(社保局|公安局|法院|检察院|派出所|公积金中心|医保局|税务局|工商局|反诈中心|网警)', ct)
    if gov_match:
        return gov_match.group(1)

    m = re.search(
        r'(?:我是|这边是|这里是)\s*([\u4e00-\u9fa5]{2,12}'
        r'(?:银行|金融|保险|证券|基金|电讯|电信|移动|联通|快递|速递|物流|'
        r'客服中心|服务中心|平台|科技|信息|投资|理财|信贷|售后|彩票|福彩|体彩|公安|'
        r'支行|分行|总部|营业部|有限公司))', ct)
    if m:
        org = m.group(1)
        org = re.sub(r'(?:客服|服务中心|售后|中心|平台|有限公司|支行|分行|营业部|总部|信贷部?)$', '', org)
        if len(org) >= 2:
            return org
    defaults = {"banking":"我行","ecommerce":"我们平台","delivery":"我司","investment":"我们机构",
                "telecom":"我们运营商","lottery":"彩票中心","kidnapping":"公安机关","identity":"信息安全中心"}
    return defaults.get(domain, "我们公司")

def _get_name(ct):
    m = re.search(r'(?:我是|叫)(小?[刘王张李陈赵孙周杨][\u4e00-\u9fa5]{0,2})', ct)
    if m: return m.group(1)
    m = re.search(r'(?:专员|经理|客服|代表)([\u4e00-\u9fa5]{2,3})', ct)
    if m: return m.group(1)
    return random.choice(["李明","王磊","张伟","刘洋","陈静"])

def _trust_intro(ct, domain, all_turns):
    org = _extract_org(ct, domain, all_turns)
    name = _get_name(ct)
    wid = random.randint(1000, 9999)
    gov_orgs = ["社保局","公安局","法院","检察院","派出所","公积金","医保局","税务局","工商局"]
    if any(g in ct for g in gov_orgs):
        enhanced = f"{org}工作人员{name}，工号{wid}。本次通话全程录音，您可拨打{org}公开办公电话或前往就近服务大厅核实。"
    else:
        intros = {
            "banking": f"{org}个人金融部客户经理{name}，工号{wid}。本次通话全程录音，并接受银保监会监管。如需核实，可拨打我行24小时客服热线转人工查询工号。",
            "ecommerce": f"{org}官方售后服务中心高级专员{name}，服务编号{wid}。您可通过平台APP内「官方客服」入口验证本次通话真实性。",
            "delivery": f"{org}客户服务部专员{name}，工号{wid}。您可登录我司官网或拨打全国统一服务热线核实我的身份。",
            "investment": f"{org}投资顾问部资深顾问{name}，执业编号1{wid:05d}。您可在中国证券业协会官网查询到我的从业资质。",
            "telecom": f"{org}客服中心高级专员{name}，工号{wid}。所有客服通话均有录音存档，您可通过运营商官方APP验证。",
            "lottery": f"{org}兑奖中心专员{name}，工号{wid}。您可通过国家彩票管理中心官网或拨打全国统一热线核实本次中奖信息。",
            "kidnapping": f"{org}刑侦支队警官{name}，警号1{wid:05d}。请您记录我的警号，可随时拨打110核实。",
            "identity": f"{org}案件专员{name}，案件编号{wid}。您可通过国家反诈中心APP或拨打96110反诈专线核实本案件。",
            "generic": f"{org}客户服务部专员{name}，工号{wid}。本次通话全程录音，您可拨打我司官网公示的客服热线核实。",
        }
        enhanced = intros.get(domain, intros["generic"])
    pat = r'(((?:我是|我这边是|我这里是|这边是|这里是))\s*[^。，]{2,30}[。，]?)'
    m = re.search(pat, ct)
    if m:
        return f"{ct[:m.start()]}{enhanced} {ct[m.end():]}".rstrip()
    return f"{enhanced} {ct}"

def _trust_verification(turns, intro_idx):
    for i in range(intro_idx + 1, len(turns)):
        if turns[i][0] == "right":
            if any(w in turns[i][1] for w in ["谁","哪位","什么","怎么","你好","嗯"]):
                turns.insert(i + 1, ("left",
                    "您放心，如果您对本次通话有任何疑虑，可以随时挂断后通过官方渠道回拨确认，我们会为您保留本次服务记录。"))
                break
    return turns

def _trust_solution(turns, domain, left_positions):
    search_start = min(2, len(left_positions))
    solution_kw = ["下载","点击","链接","APP","填写","输入","操作","验证码","提供信息","银行卡"]
    identity_kw = ["经理","专员","工号","执业编号","客服热线"]
    candidates = [p for p in left_positions[search_start:]
                  if any(kw in turns[p][1] for kw in solution_kw)
                  and not any(kw in turns[p][1] for kw in identity_kw)]
    target = candidates[0] if candidates else (left_positions[-2] if len(left_positions) >= 3 else None)
    if target is None: return turns
    stamps = {
        "banking":"该操作全程在我行经银保监会认证的安全系统内完成，数据传输采用银行级加密。",
        "ecommerce":"该流程走的是平台官方售后通道，全程在平台安全保障体系内进行。",
        "delivery":"该理赔流程走的是我司经备案的客户补偿系统，全程可追踪。",
        "investment":"该流程在我司经证监会备案的合规交易系统内完成，资金由第三方银行托管。",
        "telecom":"该业务在我司经工信部备案的服务系统内办理，全程加密传输。",
        "lottery":"该兑奖流程在经财政部认证的系统中完成，所有记录均可查验。",
        "kidnapping":"我们在依法处理此事，所有流程均严格按照公安机关办案规范执行。",
        "identity":"我们在依法核实您的信息，所有操作均严格按照个人信息保护相关规定执行。",
        "generic":"该流程在我司经认证的安全系统内完成，所有操作均有合规备案。",
    }
    stamp = stamps.get(domain, stamps["generic"])
    sp, ct = turns[target]
    turns[target] = (sp, f"{stamp} 接下来，{ct}")
    return turns

def _trust_backing(turns, domain, left_positions):
    if len(left_positions) < 2: return turns
    backings = {
        "banking":"顺便跟您确认，我行是经银保监会依法批准设立的持牌金融机构，所有业务均受《商业银行法》保护，存款还享有存款保险制度保障。",
        "ecommerce":"顺便跟您确认，我们平台是经商务部认证的电商企业，所有交易均受《电子商务法》和消费者权益保护相关法规的保障。",
        "delivery":"顺便跟您确认，我司是经邮政管理局备案的正规快递企业，所有快件均有保险保障，理赔流程严格遵循《快递暂行条例》规定。",
        "investment":"顺便跟您确认，我们是经证监会核准的持牌机构，所有投资产品均已备案，相关信息在证监会官网可查。",
        "telecom":"顺便跟您确认，我们是工信部颁发牌照的基础电信运营商，所有业务办理均严格按照《电信条例》执行。",
        "lottery":"顺便跟您确认，我们是经财政部批准的彩票发行机构，所有中奖信息均可在国家彩票管理中心官网查询验证。",
        "identity":"顺便跟您确认，我们是依法执行公务的政府工作人员，本次核查全程录音录像，所有操作均严格按照法律法规执行。",
        "kidnapping":"顺便跟您确认，我们是依法办案的公安人员，您如有疑虑可随时拨打110核实我们的警号和案件信息。",
        "generic":"顺便跟您确认，我们公司成立已超过十年，累计服务用户超过百万，所有业务均在监管部门备案，本次服务完全合规。",
    }
    phrase = backings.get(domain, backings["generic"])
    turns.insert(left_positions[-2] + 1, ("left", phrase))
    return turns

def _trust_nonfraud(turns, left_positions):
    first = left_positions[0]
    _, ct = turns[first]
    if not any(k in ct for k in ["客服","咨询","预约","预订","订餐","订","服务","餐厅","外卖","银行","快递","助手"]):
        return turns
    wid = random.randint(1000,9999)
    m = re.search(r'((?:我是|这边是|这里是)[^，。]*[，。]?)', ct)
    if m:
        turns[first] = (turns[first][0], f"{ct[:m.start()]}{m.group(1).rstrip('，。')}，工号{wid}。{ct[m.end():]}")
    else:
        turns[first] = (turns[first][0], f"{ct}（工号{wid}）")
    return turns

def apply_trust(turns, is_fraud):
    if not turns: return turns
    t = [(s,c) for s,c in turns]
    domain = detect_domain("\n".join(f"{s}: {c}" for s,c in t))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    if not lp: return t
    if not is_fraud: return _trust_nonfraud(t, lp)
    intro_idx, intro_ct = _find_intro_turn(t)
    if intro_idx >= 0: t[intro_idx] = ("left", _trust_intro(intro_ct, domain, t))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    t = _trust_verification(t, intro_idx)
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    t = _trust_solution(t, domain, lp)
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    t = _trust_backing(t, domain, lp)
    return t

# ========================= 策略2：制造紧迫感（内联核心逻辑） =========================

URGENCY_DEADLINE = {
    "banking":["需要提醒您，这个优惠利率的审批通道今天下午5点就关闭了，过了这个时间只能按标准利率申请。","我刚才确认了一下，这个低息额度仅剩今天可以受理，建议您尽快完成申请。"],
    "ecommerce":["另外提醒您，这个退款通道是限时开放的，超过48小时系统会自动关闭，届时将无法在线处理。","需要跟您说明，售后处理有时效要求的，这个订单的处理期限只剩今天了。"],
    "delivery":["另外跟您说一下，根据公司规定，理赔申请必须在包裹丢失后72小时内提交，您的包裹已经过了两天了，时间非常紧张。"],
    "investment":["需要提醒您，这个投资产品的认购期今天下午就截止了，下一期什么时候开放还不确定。","我刚才查了一下，这个额度目前只剩下最后几个名额了，估计今天之内就会满额。"],
    "telecom":["另外提醒您，这个优惠套餐的办理通道今晚12点就关闭了，明天恢复原价。"],
    "lottery":["需要提醒您，根据彩票管理条例，中奖者须在开奖之日起60个自然日内兑奖，您的兑奖期限已经非常紧迫了。"],
    "kidnapping":["情况非常紧急，每拖延一分钟风险就增加一分，必须马上处理。"],
    "identity":["您需要尽快处理，系统显示您的账户将在24小时内被冻结。"],
    "generic":["需要提醒您，这个处理是有时效限制的，超过期限系统会自动关闭。"],
}

URGENCY_SCARCITY = {
    "banking":["我刚才又查了一下，这个额度的名额确实不多了，现在已经有几十人在排队申请了。"],
    "ecommerce":["我看了一下系统，目前排队处理退款的用户非常多，如果不尽快提交，可能要排到很后面。"],
    "delivery":["理赔通道每小时只开放有限名额，现在已经有不少人在申请了。"],
    "investment":["我刚才刷新了一下系统，认购进度已经到90%了，再犹豫就真的没了。","好几个客户都在同时咨询这个产品，额度消耗得非常快。"],
    "lottery":["兑奖窗口的排号已经排到很后面了，建议您立即操作。"],
    "kidnapping":["对方情绪很不稳定，拖下去不知道会做出什么事。"],
    "identity":["冻结倒计时已经开始，现在处理还来得及，再拖就真的晚了。"],
    "generic":["我看了一下后台数据，参与的人数增长非常快，名额很快就会被抢完。"],
}

URGENCY_FINAL = {
    "banking":["好了，最后再跟您强调一下，今天下午5点是最后截止时间，过了时间我就算想帮您也帮不了了。"],
    "ecommerce":["总之请您务必在48小时内操作，否则退款通道就关闭了。"],
    "delivery":["过了今天就真的不能再申请了，系统会自动归档，到时候要走线下流程会非常麻烦。"],
    "investment":["认购今天晚上截止，建议您现在就操作，不要等到最后。"],
    "lottery":["兑奖期限真的不多了，请您务必今天就处理，明天就超过最后期限了。"],
    "identity":["冻结时间就在今天下午3点，您必须在此之前完成核实。"],
    "generic":["总之这个是有时间窗口的，请您务必抓紧，过了时间我也没办法了。"],
}

def _find_solution_turn(turns, lp):
    sol_kw = ["下载","点击","链接","APP","填写","输入","操作","提供","发送","验证码","申请","退款","理赔","转账"]
    id_kw = ["经理","专员","工号","执业编号","客服热线"]
    for p in lp[max(1,len(lp)//2):]:
        if any(k in turns[p][1] for k in sol_kw) and not any(k in turns[p][1] for k in id_kw): return p
    for p in lp[1:]:
        if any(k in turns[p][1] for k in sol_kw): return p
    return None

def _find_hesitation(turns):
    hk = ["安全","真的","确定","不太","想想","考虑","等一下","再看看","怕","担心","不太放心","怎么确定","正规","可靠"]
    return [i for i,(s,c) in enumerate(turns) if s=="right" and any(k in c for k in hk)]

def apply_urgency(turns, is_fraud):
    if not turns: return turns
    t = [(s,c) for s,c in turns]
    domain = detect_domain("\n".join(f"{s}: {c}" for s,c in t))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    if not lp: return t
    if not is_fraud:
        if any(k in t[lp[0]][1] for k in ["客服","咨询","预约","预订","订餐","订","服务","餐厅","外卖","快递"]):
            t.insert(lp[-1]+1, ("left","对了提醒您一下，这个活动到这个月底就结束了，建议尽早确认哦。"))
        return t
    sol = _find_solution_turn(t, lp)
    if sol is not None:
        pool = URGENCY_DEADLINE.get(domain, URGENCY_DEADLINE["generic"])
        t.insert(sol + 1, ("left", random.choice(pool)))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    hp = _find_hesitation(t)
    if hp:
        pool = URGENCY_SCARCITY.get(domain, URGENCY_SCARCITY["generic"])
        t.insert(hp[0] + 1, ("left", random.choice(pool)))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    final_pool = URGENCY_FINAL.get(domain)
    if final_pool and len(lp) >= 2:
        t.insert(lp[-2] + 1, ("left", random.choice(final_pool)))
    return t

# ========================= 策略3：情感操纵（内联核心逻辑） =========================

EMO_GREED = {
    "banking":["说实话，这个利率我在行里干了这么多年都很少见到，真的是难得的好政策。"],
    "ecommerce":["很多客户都通过这个通道快速拿到了退款，而且还额外获得了平台补贴。"],
    "delivery":["这次的赔偿金额是按照最高标准核算的，比平时多了不少。"],
    "investment":["不瞒您说，我自己也投了这个项目，上个月收益真的到账了。","之前有位客户投了十万，三个月就赚了两万多，他自己都不敢相信。"],
    "telecom":["这个套餐比你现在用的划算太多了，一个月能省好几十块。"],
    "lottery":["您想想，这可是五百万啊，多少人一辈子都赚不到这个数。"],
    "kidnapping":["人没事比什么都重要，钱可以再赚，人没了就什么都没了。"],
    "identity":["您的信用记录一直很好，如果因为这次被冻结，以后再想恢复就难了。"],
    "generic":["说真的，这个机会真的很难得，我自己看了都心动。"],
}

EMO_FEAR = {
    "banking":["我不是吓您，如果逾期不处理的话，系统会自动上报征信，那影响可就大了。"],
    "ecommerce":["如果不在规定时间内处理，系统会自动关闭售后通道，到时候只能走线下，手续会非常麻烦。"],
    "delivery":["如果逾期不受理，系统会自动归档，到时候再想索赔就很困难了。","之前有个客户就是处理晚了，最后只赔了一半，实在可惜。"],
    "investment":["这个认购真的不能犹豫，现在行情变化很快，今天一个价明天一个价。"],
    "lottery":["不是跟您开玩笑，逾期不兑奖的话奖金直接充公，一分都拿不到。"],
    "kidnapping":["每拖延一分钟，情况就更危险一分，真的不能再等了。"],
    "identity":["账户一旦被冻结，您的所有资金都会被锁定，到时候吃饭都成问题。"],
    "generic":["不处理的话后面会很麻烦，现在处理还来得及，再拖就真的晚了。"],
}

EMO_SYMPATHY = {
    "banking":["其实我也挺不容易的，这个月的业绩压力很大，您帮我一把，我帮您把事办好，大家都好。"],
    "ecommerce":["我们客服也挺为难的，公司给的指标完不成就要扣工资，您理解一下配合一下好吗。"],
    "delivery":["我这边也是按公司规定办事，您尽快处理了我也好交差。"],
    "investment":["我是真的觉得这个项目好才推荐给您的，您要是错过了我也替您可惜。","说实话我也就赚个辛苦费，但真心希望您能赚到钱。"],
    "lottery":["我也替您高兴啊，您好运来了挡都挡不住，赶紧兑奖吧别犹豫了。"],
    "kidnapping":["我也是受人之托，您配合一下，大家都好过。"],
    "generic":["其实我这边也挺难的，领导催得紧，您帮我个忙，我肯定尽最大努力帮您。"],
}

EMO_CARE = {
    "banking":["我是真的为您着想才打这个电话的，您想想如果没人提醒，到时候征信出问题多麻烦。","看到您的资质这么好，我第一时间就想到您了，真不希望您错过这么好的机会。"],
    "ecommerce":["我们客服主动联系您，就是怕您错过处理时间，真的是为您的权益着想。"],
    "delivery":["我第一时间通知您，就是怕您错过了理赔时效，到时候吃亏的还是您自己。"],
    "investment":["我帮客户理财这么多年，真心希望大家都能赚到钱，您的资金安全也是我最关心的。","说实话，我是把您当朋友才专门打电话来告诉您的，别人我还不一定说呢。"],
    "lottery":["我们是真心为您高兴，看到您中奖我们比您还激动，赶紧来兑奖吧。"],
    "generic":["我是真心为您好才打电话提醒您的，真的不希望您吃亏。","看到您的情况，我第一时间就想到了您，希望能帮到您。"],
}

def _find_bait_turn(turns, lp):
    bait_kw = ["机会","优惠","收益","回报","中奖","大奖","退款","补偿","便宜","免费","赠送","特别","专属","优先","赚钱","翻倍","利率低","低息","高收益"]
    id_kw = ["经理","专员","工号","执业编号"]
    for p in lp[1:]:
        if any(k in turns[p][1] for k in bait_kw) and not any(k in turns[p][1] for k in id_kw): return p
    for p in lp[1:]:
        if len(turns[p][1]) > 20: return p
    return None

def _find_refusal(turns):
    rk = ["不需要","不要了","不用了","没兴趣","不感兴趣","没时间","太忙","算了","再说","暂时不","不考虑","没钱"]
    for i,(s,c) in enumerate(turns):
        if s=="right" and any(k in c for k in rk): return i
    return None

def apply_emotional(turns, is_fraud):
    if not turns: return turns
    t = [(s,c) for s,c in turns]
    domain = detect_domain("\n".join(f"{s}: {c}" for s,c in t))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    if not lp: return t
    if not is_fraud:
        if any(k in t[lp[0]][1] for k in ["客服","咨询","预约","预订","订餐","订","服务","餐厅","外卖","快递"]):
            t.insert(lp[-1]+1, ("left","感谢您一直以来的支持，我们很珍惜与您的合作关系。"))
        return t
    bait = _find_bait_turn(t, lp)
    if bait is not None:
        pool = EMO_GREED.get(domain, EMO_GREED["generic"])
        t.insert(bait + 1, ("left", random.choice(pool)))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    hp = _find_hesitation(t)
    if hp:
        pool = EMO_FEAR.get(domain, EMO_FEAR["generic"])
        t.insert(hp[0] + 1, ("left", random.choice(pool)))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    rp = _find_refusal(t)
    if rp is not None:
        pool = EMO_SYMPATHY.get(domain, EMO_SYMPATHY["generic"])
        t.insert(rp + 1, ("left", random.choice(pool)))
    lp = [i for i,(s,_) in enumerate(t) if s=="left"]
    if lp:
        pool = EMO_CARE.get(domain, EMO_CARE["generic"])
        t.insert(lp[-1] + 1, ("left", random.choice(pool)))
    return t

# ========================= 组合策略 =========================

def apply_combined(turns, is_fraud):
    """依次叠加信任→紧迫→情感"""
    t = apply_trust(turns, is_fraud)
    t = apply_urgency(t, is_fraud)
    t = apply_emotional(t, is_fraud)
    return t

# ========================= 主流程 =========================

def main():
    project_root = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(os.path.dirname(project_root), "data")
    src_path = os.path.join(data_dir, "test.csv")

    print("=" * 60)
    print("策略4：三策略组合（Trust + Urgency + Emotional）")
    print("=" * 60)
    print(f"  数据源: {src_path}")

    df = pd.read_csv(src_path, encoding="utf-8-sig")
    if df.columns[0].startswith("\ufeff"):
        df.columns = [c.replace("\ufeff", "") for c in df.columns]
    print(f"  总样本: {len(df)}")

    is_fraud_series = df["is_fraud"].astype(str).str.upper().str.strip() == "TRUE"
    print(f"  诈骗: {is_fraud_series.sum()}, 非诈骗: {(~is_fraud_series).sum()}")

    new_texts = []
    for idx, row in df.iterrows():
        turns = parse_dialogue(row["specific_dialogue_content"])
        if turns:
            augmented = apply_combined(turns, is_fraud_series.loc[idx])
            new_texts.append(rebuild_dialogue(augmented))
        else:
            new_texts.append(row["specific_dialogue_content"])
        if (idx + 1) % 500 == 0:
            print(f"  进度: {idx + 1}/{len(df)}")

    new_df = df.copy()
    new_df["specific_dialogue_content"] = new_texts
    changed = sum(df["specific_dialogue_content"].iloc[i] != new_texts[i] for i in range(len(df)))
    orig_len = df["specific_dialogue_content"].astype(str).str.len().mean()
    aug_len = new_df["specific_dialogue_content"].astype(str).str.len().mean()
    print(f"  改写: {changed}/{len(df)} 条")
    print(f"  长度: {orig_len:.0f} → {aug_len:.0f} (+{aug_len-orig_len:.0f}字)")

    out_path = os.path.join(data_dir, "test_combined.csv")
    new_df.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"  输出: {out_path}")

    # ==================== 生成对比 Excel ====================
    print(f"\n{'=' * 60}")
    print("生成代表性对比 Excel（含颜色标记）……")
    print(f"{'=' * 60}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        print("  openpyxl 未安装，尝试安装中……")
        os.system("pip install openpyxl -q")
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # 读入各策略结果
    aug_trust = pd.read_csv(os.path.join(data_dir, "test_trust.csv"), encoding="utf-8-sig")
    aug_urgency = pd.read_csv(os.path.join(data_dir, "test_urgency.csv"), encoding="utf-8-sig")
    aug_emotional = pd.read_csv(os.path.join(data_dir, "test_emotional.csv"), encoding="utf-8-sig")
    aug_combined = new_df

    # 选取代表性样本：3条非诈骗 + 7种诈骗类型各1 + 额外2条
    selected = []
    nf = df[~is_fraud_series]
    if len(nf) >= 3:
        selected.extend(nf.sample(3, random_state=42).index.tolist())
    for ft in ["客服诈骗","银行诈骗","投资诈骗","钓鱼诈骗","彩票诈骗","绑架诈骗","身份盗窃"]:
        subset = df[df["fraud_type"] == ft]
        if len(subset) > 0:
            selected.append(subset.sample(1, random_state=42).index[0])
    extra = df[is_fraud_series & ~df.index.isin(selected)]
    if len(extra) >= 2:
        selected.extend(extra.sample(2, random_state=42).index.tolist())

    print(f"  选取了 {len(selected)} 条代表性样本")

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    normal_font = Font(name="微软雅黑", size=9)
    bold_font = Font(name="微软雅黑", size=9, bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    # 差异标记颜色（浅色背景）
    highlight_fills = {
        "trust": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),     # 浅黄
        "urgency": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),   # 浅橙
        "emotional": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), # 浅绿
        "combined": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # 浅蓝
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "策略对比"

    # 列：sample_id | 标签信息 | 原始 | 信任 | 紧迫 | 情感 | 组合
    headers = ["样本编号", "是否诈骗", "诈骗类型", "原始对话",
               "策略1-建立信任", "策略2-制造紧迫感", "策略3-情感操纵", "策略4-三策略组合"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin_border

    # 列宽
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    for col_letter in ["D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 45

    row_num = 2
    for sample_id, idx in enumerate(selected, 1):
        row_data = df.loc[idx]
        orig_text = str(row_data["specific_dialogue_content"])
        is_f = str(row_data["is_fraud"])
        fraud_t = str(row_data.get("fraud_type", ""))
        if fraud_t == "nan":
            fraud_t = "—"

        trust_text = str(aug_trust.loc[idx, "specific_dialogue_content"])
        urgency_text = str(aug_urgency.loc[idx, "specific_dialogue_content"])
        emotional_text = str(aug_emotional.loc[idx, "specific_dialogue_content"])
        combined_text = str(aug_combined.loc[idx, "specific_dialogue_content"])

        texts = [
            (orig_text, None),           # 原始不标色
            (trust_text, "trust"),       # 信任策略
            (urgency_text, "urgency"),   # 紧迫策略
            (emotional_text, "emotional"), # 情感策略
            (combined_text, "combined"),  # 组合策略
        ]

        # 前3列：编号 + 标签
        ws.cell(row=row_num, column=1, value=f"样本{sample_id}").font = bold_font
        ws.cell(row=row_num, column=2, value=is_f).font = normal_font
        ws.cell(row=row_num, column=3, value=fraud_t).font = normal_font

        # 对话列
        for col_offset, (txt, strategy_key) in enumerate(texts):
            cell = ws.cell(row=row_num, column=4 + col_offset, value=txt)
            cell.font = normal_font
            cell.alignment = wrap

            # 颜色标记：与原数据不同则标色
            if strategy_key is not None and txt != orig_text:
                cell.fill = highlight_fills[strategy_key]

        # 边框
        for c in range(1, 9):
            ws.cell(row=row_num, column=c).border = thin_border

        # 行高自适应（大致估算）
        max_len = max(len(t) for t in [orig_text, trust_text, urgency_text, emotional_text, combined_text])
        ws.row_dimensions[row_num].height = max(60, min(300, max_len // 2))

        row_num += 1

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加图例说明 sheet
    ws2 = wb.create_sheet("图例说明")
    ws2.cell(row=1, column=1, value="颜色标记说明").font = Font(name="微软雅黑", bold=True, size=12)
    legends = [
        ("浅黄色背景", "策略1-建立信任：与原数据不同的部分"),
        ("浅橙色背景", "策略2-制造紧迫感：与原数据不同的部分"),
        ("浅绿色背景", "策略3-情感操纵：与原数据不同的部分"),
        ("浅蓝色背景", "策略4-三策略组合：与原数据不同的部分"),
        ("无颜色", "原始对话数据"),
    ]
    for r, (color, desc) in enumerate(legends, 3):
        ws2.cell(row=r, column=1, value=color).font = Font(name="微软雅黑", size=10)
        ws2.cell(row=r, column=2, value=desc).font = Font(name="微软雅黑", size=10)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 50

    xlsx_path = os.path.join(data_dir, "test_comparison.xlsx")
    wb.save(xlsx_path)
    print(f"  已保存: {xlsx_path}")

    print(f"\n{'=' * 60}")
    print(f"完成！")
    print(f"  组合测试集: {out_path}")
    print(f"  对比 Excel: {xlsx_path}")
    print(f"{'=' * 60}")

if __name__ == "__main__":
    main()
